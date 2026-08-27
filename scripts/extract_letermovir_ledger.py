#!/usr/bin/env python3
"""Regenerate benchmarks/letermovir/ledger.yaml from the published LCA workbook.

Source paper
------------
Sorgenfrei, Werz, et al., "Integrated Life Cycle Assessment Guides
Sustainability in Synthesis: Antiviral Letermovir as a Case Study",
J. Am. Chem. Soc. 2025, 147, 40944-40957.
DOI: 10.1021/jacs.5c14470. Open access via PubMed Central: PMC12593353.

The workbook this script reads (ja5c14470_si_002.xlsx, "the LCA workbook")
is Supporting Information for that article. It was retrieved from Europe
PMC's supplementary-files endpoint for the PMC id above:

    https://www.ebi.ac.uk/europepmc/webservices/rest/PMC12593353/supplementaryFiles

(that endpoint returns a zip containing ja5c14470_si_001.pdf, the SI text,
and ja5c14470_si_002.xlsx, the workbook this script consumes).

Licensing -- read before changing what this script touches
-------------------------------------------------------------
The workbook's own GWP figures were computed by the paper's authors with
**ecoinvent 3.10**, a commercially licensed background LCI database that
this project may never redistribute (see benchmarks/README.md and
CONTRIBUTING.md). Conveniently, every `GWP kgCO2-eq[/kg]` cell in the copy
of the workbook shipped with the article is cached as 0 (its macro was not
evaluated on save), so there is nothing numeric to accidentally leak from
those columns -- but this script does not read them regardless, on
principle, and it never reads any cell from the "A. Databases" sheet
except (by a human, not this script) to confirm which column holds bare
material names. If you ever see a non-zero value under a GWP heading in
this workbook, leave it exactly where it is; it must never be copied into
this repository.

What this script *does* extract, all of it public-record, non-ecoinvent
data: material names, the "Mass Stoich" column (mass already back-computed
by the paper's authors to their functional unit of 1 kg letermovir), step
names/descriptions, and each step's published yield (recorded only as a
YAML comment -- see NOTE below on why yield is not applied to mass).

Sheet layout this script depends on (both are "main route" sheets -- see
module docstring end for what is deliberately left unparsed):

  "2.1. Merck LCA FU"     (route "merck")
      columns: A name | B given-stoich | C unit | D Mass Stoich (kg) |
               E unit | F GWP/kg (ignored) | G GWP (ignored) |
               H/I a 4-row per-step "legend" table (Starting Materials /
               Reagents / Catalysts / Solvent), each row a formula that
               SUMs specific G-column cells -- this is the *authors'
               own* fine-grained role classification for the "Reactants"
               block, expressed as spreadsheet formulas rather than a
               plain heading. This script parses those formula strings
               (never their computed values) to recover which material
               row is a starting material / reagent / catalyst. No
               chemical judgement of ours is involved.
      No explicit per-step "Yield" row exists on this sheet.

  "1.1 LCA This Work FU"  (route "denovo")
      columns: A name | B given-stoich | C unit | D moles (optional) |
               E unit | F Mass Stoich (kg) | G unit | H GWP/kg (ignored) |
               I GWP (ignored)
      Has an explicit "Yield" row per step (steps 1-6; step 7 has none).
      This sheet carries no H/I legend, so role classification here is
      coarser: everything under the "Reactants" heading is `reactant`
      except a row literally named "Catalyst" (the sheet's own bespoke
      placeholder name for an undisclosed chiral catalyst), which is
      `catalyst`. See SOURCES.md for the full mapping and its rationale.

Both sheets: material rows are grouped under plain A-column headings
"Reactants" / "Solvents" / "Product"; a "Product" block is the step's
output (already consumed as the following step's "Reactants" entry) and
is never itself emitted as an input.

NOTE on yield=1.0
------------------
The "Mass Stoich" values are *already* the authors' functional-unit-scaled
masses (1 kg letermovir out). Applying carbonroute's own step-yield mass
conversion on top of an already-FU-scaled mass would double-count the
yield loss. So every step in the generated ledger is written with
`yield: 1.0`, and the workbook's own published yield (when present) is
recorded only as a YAML comment for a reader to see.

CAS resolution
--------------
Material names are resolved to CAS registry numbers via PubChem PUG REST
(name -> CID -> synonyms), rate-limited to <=5 req/s, with a local JSON
cache (see --cas-cache) so re-runs do not re-hit the network. A bare
compound number/code (e.g. "15", "11a", "8a (8)") or an explicitly bespoke
placeholder (e.g. "Catalyst") is never sent to PubChem -- it is recognised
by pattern and recorded as cas: null directly. Anything that does not
resolve to exactly one CID, or whose best CAS-shaped synonym fails the
mod-10 check digit, is also recorded as cas: null. A null CAS is treated
as a correct, honest answer -- never as a reason to guess.

Usage
-----
    PYTHONPATH=src python3 scripts/extract_letermovir_ledger.py \\
        /path/to/ja5c14470_si_002.xlsx \\
        --out benchmarks/letermovir/ledger.yaml \\
        --cas-cache /tmp/letermovir_cas_cache.json

Add --offline to skip all network access and resolve strictchemical names
only from --cas-cache (anything not already cached becomes cas: null).
Add --log <path> to also write the full resolution log as JSON (for
transcription into SOURCES.md); by default the log is only printed.

This script is deterministic given a fixed workbook and a fixed CAS
cache / PubChem state: it never invents a mass, a material, or a role. If
a section is a structural surprise it cannot classify with confidence, it
raises rather than guessing.
"""
from __future__ import annotations

import argparse
import json
import re
import sys
import time
import urllib.error
import urllib.parse
import urllib.request
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "src"))
from carbonroute.schema import cas_checksum_ok  # noqa: E402

MERCK_SHEET = "2.1. Merck LCA FU"
DENOVO_SHEET = "1.1 LCA This Work FU"

ROLE_LEGEND_MAP = {
    "Starting Materials": "reactant",
    "Reagents": "reagent",
    "Catalysts": "catalyst",
    "Solvent": "solvent",
}

# Bare compound numbers/codes used by the workbook to refer to
# intermediates carried from one step to the next (e.g. "15", "11a",
# "8a (8)"), and the sheets' own placeholder name for an undisclosed
# catalyst. Neither is a real chemical name, so neither is ever sent to
# PubChem; both are recorded as cas: null by construction.
_COMPOUND_LABEL_RE = re.compile(r"^\d+[a-zA-Z]?(\s*\(\d+\))?$")


def is_bespoke_or_intermediate(name: str) -> tuple[bool, str]:
    s = name.strip()
    if _COMPOUND_LABEL_RE.match(s):
        return True, "bare compound number/label (intermediate carried between steps)"
    if s.lower() == "catalyst":
        return True, "bespoke/undisclosed catalyst placeholder name"
    return False, ""


# ---------------------------------------------------------------------------
# Workbook parsing
# ---------------------------------------------------------------------------


def _cell(ws, row: int, col: int):
    return ws.cell(row=row, column=col).value


def _parse_g_row_refs(formula) -> set[int]:
    """Extract row numbers referenced in column G by a legend formula.

    Handles '=G8', '=SUM(G6,G7)', '=SUM(G10:G11)', and combinations. Never
    evaluates the formula -- only reads which G-cells it points at.
    """
    if not isinstance(formula, str) or not formula.startswith("="):
        return set()
    body = formula[1:]
    rows: set[int] = set()
    for m in re.finditer(r"G(\d+):G(\d+)", body):
        a, b = int(m.group(1)), int(m.group(2))
        rows.update(range(a, b + 1))
    for m in re.finditer(r"G(\d+)", body):
        rows.add(int(m.group(1)))
    return rows


class ParsedMaterial:
    def __init__(self, row: int, name: str, mass_kg: float, role: str):
        self.row = row
        self.name = name
        self.mass_kg = mass_kg
        self.role = role


class ParsedStep:
    def __init__(self, step_id: int | str, desc: str, header_row: int,
                 materials: list[ParsedMaterial], published_yield: float | None,
                 yield_row: int | None):
        self.step_id = step_id
        self.desc = desc
        self.header_row = header_row
        self.materials = materials
        self.published_yield = published_yield
        self.yield_row = yield_row


def _find_step_blocks(ws_values) -> list[tuple[int | str, str, int, int]]:
    """Return (step_id, desc, header_row, block_end_row_exclusive) tuples."""
    max_row = ws_values.max_row
    blocks = []
    starts = []
    for row in range(1, max_row + 1):
        a = _cell(ws_values, row, 1)
        if isinstance(a, str) and a.strip().lower().startswith("step"):
            m = re.search(r"\d+", a)
            step_id = int(m.group()) if m else a.strip()
            desc = _cell(ws_values, row, 2)
            starts.append((step_id, desc, row))
    for i, (step_id, desc, row) in enumerate(starts):
        end = starts[i + 1][2] if i + 1 < len(starts) else max_row + 1
        blocks.append((step_id, desc, row, end))
    return blocks


def parse_merck_sheet(ws_values, ws_formulas) -> list[ParsedStep]:
    mass_col = 4  # D: Mass Stoich
    steps: list[ParsedStep] = []
    for step_id, desc, header_row, block_end in _find_step_blocks(ws_values):
        sec = {}
        for r in range(header_row + 1, block_end):
            av = _cell(ws_values, r, 1)
            if isinstance(av, str) and av.strip() in ("Reactants", "Solvents", "Product"):
                sec[av.strip()] = r
        react_row = sec.get("Reactants")
        solv_row = sec.get("Solvents")
        prod_row = sec.get("Product")
        if react_row is None or solv_row is None or prod_row is None:
            raise ValueError(
                f"{MERCK_SHEET} step {step_id}: expected Reactants/Solvents/Product "
                f"sections, found {sec!r} -- refusing to guess, stopping."
            )

        # The authors' own role legend: 4 rows starting at the Reactants
        # header, each an (H label, I formula) pair. The formula's G-cell
        # references are the material rows that belong to that label.
        row_role: dict[int, str] = {}
        for lr in range(react_row, react_row + 4):
            label = _cell(ws_formulas, lr, 8)
            formula = _cell(ws_formulas, lr, 9)
            if label in ROLE_LEGEND_MAP:
                for gr in _parse_g_row_refs(formula):
                    row_role[gr] = ROLE_LEGEND_MAP[label]

        materials = []
        for r in range(react_row + 1, solv_row):
            name = _cell(ws_values, r, 1)
            mass = _cell(ws_values, r, mass_col)
            if name is None or mass is None:
                continue
            role = row_role.get(r)
            if role is None:
                raise ValueError(
                    f"{MERCK_SHEET} step {step_id} row {r} ({name!r}): not covered "
                    "by the H/I role legend -- refusing to guess a role."
                )
            materials.append(ParsedMaterial(r, str(name), float(mass), role))
        for r in range(solv_row + 1, prod_row):
            name = _cell(ws_values, r, 1)
            mass = _cell(ws_values, r, mass_col)
            if name is None or mass is None:
                continue
            materials.append(ParsedMaterial(r, str(name), float(mass), "solvent"))

        steps.append(ParsedStep(step_id, desc, header_row, materials, None, None))
    return steps


def parse_denovo_sheet(ws_values) -> list[ParsedStep]:
    mass_col = 6  # F: Mass Stoich
    steps: list[ParsedStep] = []
    for step_id, desc, header_row, block_end in _find_step_blocks(ws_values):
        sec = {}
        yield_row = None
        published_yield = None
        for r in range(header_row + 1, block_end):
            av = _cell(ws_values, r, 1)
            if isinstance(av, str) and av.strip() in ("Reactants", "Solvents", "Product"):
                sec[av.strip()] = r
            elif isinstance(av, str) and av.strip() == "Yield":
                yield_row = r
                published_yield = _cell(ws_values, r, 2)
        react_row = sec.get("Reactants")
        solv_row = sec.get("Solvents")
        prod_row = sec.get("Product")
        if react_row is None or solv_row is None or prod_row is None:
            raise ValueError(
                f"{DENOVO_SHEET} step {step_id}: expected Reactants/Solvents/Product "
                f"sections, found {sec!r} -- refusing to guess, stopping."
            )

        materials = []
        for r in range(react_row + 1, solv_row):
            name = _cell(ws_values, r, 1)
            mass = _cell(ws_values, r, mass_col)
            if name is None or mass is None:
                continue
            if isinstance(name, str) and name.strip() == "Yield":
                continue
            role = "catalyst" if isinstance(name, str) and name.strip().lower() == "catalyst" else "reactant"
            materials.append(ParsedMaterial(r, str(name), float(mass), role))
        for r in range(solv_row + 1, prod_row):
            name = _cell(ws_values, r, 1)
            mass = _cell(ws_values, r, mass_col)
            if name is None or mass is None:
                continue
            if isinstance(name, str) and name.strip() == "Yield":
                continue
            materials.append(ParsedMaterial(r, str(name), float(mass), "solvent"))

        steps.append(ParsedStep(step_id, desc, header_row, materials, published_yield, yield_row))
    return steps


# ---------------------------------------------------------------------------
# CAS resolution (PubChem PUG REST)
# ---------------------------------------------------------------------------

PUBCHEM_BASE = "https://pubchem.ncbi.nlm.nih.gov/rest/pug"
_CAS_SHAPE_RE = re.compile(r"^\d{2,7}-\d{2}-\d$")


def _normalize_query_candidates(name: str) -> list[str]:
    s = name.strip()
    cands = [s]
    # strip a leading molarity descriptor: "1 M NaOH" -> "NaOH"
    s2 = re.sub(r"^\d+(\.\d+)?\s*M\s+", "", s, flags=re.I)
    if s2 != s and s2:
        cands.append(s2)
    # strip a leading "aq." descriptor: "aq. NaHCO3" -> "NaHCO3"
    s3 = re.sub(r"^aq\.\s*", "", s, flags=re.I)
    if s3 != s and s3:
        cands.append(s3)
    s4 = re.sub(r"^aq\.\s*", "", s2, flags=re.I)
    if s4 not in cands and s4 != s2:
        cands.append(s4)
    # close up a stray space after a hyphen: "2- picoline" -> "2-picoline"
    s5 = re.sub(r"-\s+", "-", s)
    if s5 not in cands:
        cands.append(s5)
    # dedupe, preserve order
    seen = set()
    out = []
    for c in cands:
        c = c.strip()
        if c and c not in seen:
            seen.add(c)
            out.append(c)
    return out


class RateLimiter:
    def __init__(self, max_per_sec: float = 5.0):
        self.min_interval = 1.0 / max_per_sec
        self._last = 0.0

    def wait(self):
        now = time.monotonic()
        dt = now - self._last
        if dt < self.min_interval:
            time.sleep(self.min_interval - dt)
        self._last = time.monotonic()


def _http_get_json(url: str, limiter: RateLimiter, timeout: int = 30):
    limiter.wait()
    req = urllib.request.Request(
        url, headers={"User-Agent": "carbonroute-letermovir-ledger/0.1 (benchmark extraction)"}
    )
    with urllib.request.urlopen(req, timeout=timeout) as resp:
        return json.loads(resp.read().decode("utf-8"))


def resolve_cas_for_name(
    name: str, cache: dict, limiter: RateLimiter, offline: bool, log: list
) -> str | None:
    """Return a validated CAS string or None, updating `cache` and `log`."""
    skip, reason = is_bespoke_or_intermediate(name)
    if skip:
        log.append({"name": name, "cas": None, "status": "skipped", "reason": reason})
        return None

    if name in cache:
        entry = cache[name]
        log.append({"name": name, "cas": entry.get("cas"), "status": "cache", "reason": entry.get("reason", "")})
        return entry.get("cas")

    if offline:
        log.append({"name": name, "cas": None, "status": "offline-uncached",
                     "reason": "not in --cas-cache and --offline was set"})
        cache[name] = {"cas": None, "reason": "offline-uncached"}
        return None

    resolved_cas = None
    reason = ""
    for candidate in _normalize_query_candidates(name):
        try:
            q = urllib.parse.quote(candidate, safe="")
            cid_doc = _http_get_json(f"{PUBCHEM_BASE}/compound/name/{q}/cids/JSON", limiter)
            cids = cid_doc.get("IdentifierList", {}).get("CID", [])
        except urllib.error.HTTPError as e:
            if e.code == 404:
                continue
            reason = f"HTTP {e.code} on CID lookup for {candidate!r}"
            continue
        except Exception as e:  # noqa: BLE001
            reason = f"error on CID lookup for {candidate!r}: {e}"
            continue

        if len(cids) != 1:
            reason = f"{candidate!r} -> {len(cids)} CIDs (need exactly 1)"
            continue

        cid = cids[0]
        try:
            syn_doc = _http_get_json(f"{PUBCHEM_BASE}/compound/cid/{cid}/synonyms/JSON", limiter)
            synonyms = syn_doc["InformationList"]["Information"][0].get("Synonym", [])
        except Exception as e:  # noqa: BLE001
            reason = f"CID {cid}: synonym lookup failed: {e}"
            continue

        cas_candidates = [s for s in synonyms if _CAS_SHAPE_RE.match(s)]
        if not cas_candidates:
            reason = f"CID {cid}: no CAS-shaped synonym"
            continue
        best = cas_candidates[0]
        if not cas_checksum_ok(best):
            reason = f"CID {cid}: synonym {best!r} failed CAS check digit"
            continue
        resolved_cas = best
        reason = f"matched via query {candidate!r} -> CID {cid} -> CAS {best}"
        break

    cache[name] = {"cas": resolved_cas, "reason": reason}
    log.append({"name": name, "cas": resolved_cas, "status": "resolved" if resolved_cas else "unresolved",
                 "reason": reason})
    return resolved_cas


# ---------------------------------------------------------------------------
# YAML emission (hand-rolled: no third-party yaml dependency assumed beyond
# PyYAML which is available, but we want inline comments per-line, which
# PyYAML's dumper does not support -- so the ledger is emitted directly.)
# ---------------------------------------------------------------------------


def _yaml_str(s: str) -> str:
    needs_quote = (
        s == ""
        or s[0] in "!&*-?|>%@`\"'#,[]{}:"
        or ": " in s
        or s.strip() != s
        or s.lower() in ("null", "true", "false", "yes", "no", "~")
        or re.match(r"^[\d.\-]+$", s)
    )
    if needs_quote:
        return '"' + s.replace("\\", "\\\\").replace('"', '\\"') + '"'
    return s


def emit_ledger_yaml(routes: dict[str, list[ParsedStep]], cas_of: dict[str, str | None],
                      source_note: str) -> str:
    lines = []
    lines.append("# Letermovir cradle-to-gate benchmark ledger (carbonroute B2).")
    lines.append("#")
    lines.append("# Generated by scripts/extract_letermovir_ledger.py from the published LCA")
    lines.append("# workbook (Sorgenfrei et al., J. Am. Chem. Soc. 2025, 147, 40944-40957,")
    lines.append("# doi:10.1021/jacs.5c14470, PMC12593353, SI file ja5c14470_si_002.xlsx).")
    lines.append("# Full provenance, role mapping, and CAS resolution log: SOURCES.md in this")
    lines.append("# directory. DO NOT hand-edit masses here -- re-run the extraction script.")
    lines.append("#")
    lines.append("# yield: 1.0 on every step is deliberate, not a placeholder: the workbook's")
    lines.append("# 'Mass Stoich' masses are already back-calculated by the paper's authors to")
    lines.append("# the functional unit (1 kg letermovir out). Applying carbonroute's own yield")
    lines.append("# conversion on top would double-count the yield loss. Each step's own")
    lines.append("# *published* yield (where the sheet gives one) is noted in a comment next to")
    lines.append("# that step instead, purely for a reader's reference.")
    lines.append("#")
    lines.append("# electricity_kWh: 0.0 on every step is also deliberate: the published")
    lines.append("# inventory in this workbook is material-only. Process energy is out of scope")
    lines.append("# for this benchmark; no energy figure has been invented or estimated.")
    lines.append('schema_version: "0.1"')
    lines.append("")
    lines.append("assumptions:")
    lines.append("  functional_unit: {mass_kg: 1.0, basis: product}")
    lines.append("  boundary: cradle-to-gate")
    lines.append("  # No electricity is charged anywhere in this benchmark (see note above), so")
    lines.append("  # the grid factor is a documented zero, not an uncited placeholder.")
    lines.append("  grid_factor:")
    lines.append('    id: "none-charged"')
    lines.append("    value_kgCO2e_per_kWh: 0.0")
    lines.append('    source: "This benchmark charges no electricity_kWh anywhere (material-only')
    lines.append('      published inventory, see benchmarks/letermovir/SOURCES.md); the grid')
    lines.append('      factor is unused by construction and is set to 0 rather than left to an')
    lines.append('      analyst-declared placeholder."')
    lines.append("    uncertainty_class: assumption")
    lines.append("  # The paper itself reports GWP100a per IPCC 2021 (AR6); recorded here as the")
    lines.append("  # nominal method label. See SOURCES.md -- this benchmark's own numbers are")
    lines.append("  # computed from public factor tables, not from the paper's ecoinvent run, so")
    lines.append("  # this label describes intent/comparability, not a reproduction of the paper's")
    lines.append("  # own GWP methodology chain.")
    lines.append("  gwp_method: {name: IPCC-AR6, horizon_years: 100, feedbacks: false}")
    lines.append("  solvent_recovery_default: 0.0")
    lines.append("  waste_treatment: excluded")
    lines.append("")
    lines.append("routes:")

    labels = {"merck": "Merck route (Sorgenfrei et al. 2025, sheet '2.1. Merck LCA FU')",
              "denovo": "De novo / academic route (Sorgenfrei et al. 2025, sheet '1.1 LCA This Work FU')"}

    for route_key, steps in routes.items():
        lines.append(f"  {route_key}:")
        lines.append(f"    label: {_yaml_str(labels[route_key])}")
        lines.append("    steps:")
        for st in steps:
            lines.append(f"      - id: {st.step_id}")
            lines.append(f"        # {st.desc}")
            if st.published_yield is not None:
                lines.append(
                    f"        # Published step yield (workbook 'Yield' row, {DENOVO_SHEET!r} "
                    f"row {st.yield_row}): {st.published_yield}"
                )
            else:
                lines.append(
                    "        # Published step yield: not given as a separate figure on this "
                    "sheet (see SOURCES.md)."
                )
            lines.append(
                "        # yield fixed at 1.0: mass_kg values below are already scaled to the"
            )
            lines.append(
                "        # functional unit; see the ledger-level comment above."
            )
            lines.append("        yield: 1.0")
            lines.append("        inputs:")
            for mat in st.materials:
                cas = cas_of.get(mat.name)
                cas_repr = f'"{cas}"' if cas else "null"
                lines.append(
                    f"          - {{name: {_yaml_str(mat.name)}, cas: {cas_repr}, "
                    f"mass_kg: {mat.mass_kg!r}, role: {mat.role}}}"
                )
            lines.append("        # No process energy in the published inventory (material-only).")
            lines.append("        electricity_kWh: 0.0")
        lines.append("")
    return "\n".join(lines).rstrip() + "\n"


# ---------------------------------------------------------------------------
# main
# ---------------------------------------------------------------------------


def main() -> None:
    ap = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("workbook", type=Path, help="path to ja5c14470_si_002.xlsx")
    ap.add_argument("--out", type=Path, required=True, help="output ledger.yaml path")
    ap.add_argument(
        "--cas-cache",
        type=Path,
        default=Path(__file__).resolve().parents[1] / "data" / "raw" / "letermovir_cas_cache.json",
        help="JSON cache of name->CAS resolutions (read+write). Defaults to a durable, "
        "committed path under data/raw/, so a plain re-run with no flags at all replays "
        "from the frozen snapshot instead of silently trying PubChem again.",
    )
    ap.add_argument("--offline", action="store_true",
                     help="never contact PubChem; unresolved names not already in --cas-cache "
                          "become cas: null")
    ap.add_argument("--log", type=Path, default=None,
                     help="optional path to also write the full CAS resolution log as JSON")
    args = ap.parse_args()

    wb_values = openpyxl.load_workbook(args.workbook, data_only=True)
    wb_formulas = openpyxl.load_workbook(args.workbook, data_only=False)

    merck_steps = parse_merck_sheet(wb_values[MERCK_SHEET], wb_formulas[MERCK_SHEET])
    denovo_steps = parse_denovo_sheet(wb_values[DENOVO_SHEET])

    if len(merck_steps) != 7 or len(denovo_steps) != 7:
        raise ValueError(
            f"expected exactly 7 main-route steps per sheet, got merck={len(merck_steps)} "
            f"denovo={len(denovo_steps)} -- stopping rather than emitting a partial ledger "
            "silently."
        )

    all_names = sorted({m.name for steps in (merck_steps, denovo_steps) for st in steps for m in st.materials})

    cache: dict = {}
    if args.cas_cache and args.cas_cache.exists():
        cache = json.loads(args.cas_cache.read_text())

    limiter = RateLimiter(max_per_sec=5.0)
    log: list = []
    cas_of: dict[str, str | None] = {}
    for name in all_names:
        cas_of[name] = resolve_cas_for_name(name, cache, limiter, args.offline, log)

    if args.cas_cache:
        args.cas_cache.parent.mkdir(parents=True, exist_ok=True)
        args.cas_cache.write_text(json.dumps(cache, indent=2, sort_keys=True))

    if args.log:
        args.log.parent.mkdir(parents=True, exist_ok=True)
        args.log.write_text(json.dumps(log, indent=2))

    resolved = sum(1 for e in log if e["status"] in ("resolved", "cache") and e["cas"])
    print(f"CAS resolution: {resolved}/{len(all_names)} names resolved to a validated CAS.",
          file=sys.stderr)
    for entry in log:
        marker = "OK  " if entry["cas"] else "NULL"
        print(f"  [{marker}] {entry['name']!r:55} status={entry['status']:10} "
              f"cas={entry['cas']!r} reason={entry['reason']}", file=sys.stderr)

    routes = {"merck": merck_steps, "denovo": denovo_steps}
    ledger_text = emit_ledger_yaml(routes, cas_of, source_note="")

    args.out.parent.mkdir(parents=True, exist_ok=True)
    args.out.write_text(ledger_text)
    print(f"wrote {args.out}", file=sys.stderr)


if __name__ == "__main__":
    main()
